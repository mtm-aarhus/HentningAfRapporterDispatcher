from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from OpenOrchestrator.database.queues import QueueElement
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
import os
from openpyxl import load_workbook
import json 
def process(orchestrator_connection: OrchestratorConnection, queue_element: QueueElement | None = None) -> None:

    #Robotpassword
    RobotCredential = orchestrator_connection.get_credential("Robot365User")
    RobotUsername = RobotCredential.username
    RobotPassword = RobotCredential.password

    #Sharepoint
    API_url = orchestrator_connection.get_constant("AarhusKommuneSharePoint").value

    #Connecting to sharepoint
    credentials = UserCredential(RobotUsername, RobotPassword)
    ctx = ClientContext(API_url + "/Teams/tea-teamsite11819").with_credentials(credentials)
    web = ctx.web
    ctx.load(web)
    ctx.execute_query()

    # SharePoint site and parent folder URL
    PARENT_FOLDER_URL = "/Teams/tea-teamsite11819/Delte Dokumenter/OPUSrobottest/ManglendeTidsregistreringDispatcher.xlsx"
    file_name = PARENT_FOLDER_URL.split("/")[-1]
    download_path = os.path.join(os.getcwd(), file_name)

    # Download the file to the specified path
    with open(download_path, "wb") as local_file:
        file = ctx.web.get_file_by_server_relative_path(PARENT_FOLDER_URL).download(local_file).execute_query()

    # Load the workbook using the file path
    workbook = load_workbook(filename=download_path)

    # Access the workbook
    ark1 = workbook["ark1"]
    ark1 = workbook.active
    row_count = ark1.max_row

    queue_items = []

    if row_count > 0:
        for row_idx in range(2, ark1.max_row + 1):  # Assuming the first row is a header
            # Extract data for the queue element
            row_data = {
                "Navn": ark1[f"A{row_idx}"].value,  # Replace with actual column names
                "QueueName": ark1[f"B{row_idx}"].value,  # Adjust column references as needed
                "Sti": ark1[f"C{row_idx}"].value,
                "SharePointMappeLink": ark1[f"D{row_idx}"].value
            }

            # Prepare queue item with SpecificContent and Reference
            queue_items.append({
                "SpecificContent": row_data,
                "Reference": ark1[f"A{row_idx}"].value  # Assuming column A provides a unique reference
            })

        # Prepare references and data for the bulk creation function
        references = tuple(item["Reference"] for item in queue_items)  # Extract references as a tuple
        data = tuple(json.dumps(item["SpecificContent"]) for item in queue_items)  # Convert SpecificContent to JSON strings

        # Bulk add queue items to OpenOrchestrator
        queue_name = "HentningAfRapporterQueue" 
        try:
            orchestrator_connection.bulk_create_queue_elements(queue_name, references, data, created_by="HentningAfRapporterBob")
            orchestrator_connection.log_info(f"Successfully added {len(queue_items)} items to the queue.")
        except Exception as e:
            print(f"An error occurred while adding items to the queue: {str(e)}")
    else:
        orchestrator_connection.log_info("Ingen bogmærker")

    if os.path.exists(file_name):
        os.remove(file_name)